"""
Build a formatted multi-sheet Excel report from asx.db.

Output: reports/asx-market-report.xlsx
Sheets: Sector Returns, Top Performers, Most Volatile, Max Drawdown,
        Sector Turnover, Companies.
"""

import os
import sqlite3

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(__file__)
DB = os.path.join(HERE, "..", "asx.db")
OUT = os.path.join(HERE, "..", "reports", "asx-market-report.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

SHEETS = {
    "Sector Returns": """
        WITH b AS (SELECT code,
              FIRST_VALUE(adj_close) OVER w AS f, LAST_VALUE(adj_close) OVER w AS l
            FROM daily_prices
            WINDOW w AS (PARTITION BY code ORDER BY trade_date
                         ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)),
        r AS (SELECT DISTINCT code, 100.0*(l-f)/f AS ret FROM b)
        SELECT s.sector_name AS Sector, COUNT(*) AS Companies,
               ROUND(AVG(r.ret),2) AS [Avg 1Y Return %]
        FROM r JOIN companies c ON c.code=r.code JOIN sectors s ON s.sector_id=c.sector_id
        GROUP BY s.sector_name ORDER BY 3 DESC""",
    "Top Performers": """
        WITH b AS (SELECT code,
              FIRST_VALUE(adj_close) OVER w AS f, LAST_VALUE(adj_close) OVER w AS l
            FROM daily_prices
            WINDOW w AS (PARTITION BY code ORDER BY trade_date
                         ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)),
        r AS (SELECT DISTINCT code, 100.0*(l-f)/f AS ret FROM b)
        SELECT c.code AS Code, c.company_name AS Company, s.sector_name AS Sector,
               ROUND(r.ret,1) AS [1Y Return %]
        FROM r JOIN companies c ON c.code=r.code JOIN sectors s ON s.sector_id=c.sector_id
        ORDER BY r.ret DESC LIMIT 25""",
    "Most Volatile": """
        WITH d AS (SELECT code, adj_close,
              LAG(adj_close) OVER (PARTITION BY code ORDER BY trade_date) AS p FROM daily_prices),
        r AS (SELECT code,(adj_close-p)/p AS x FROM d WHERE p IS NOT NULL)
        SELECT code AS Code, COUNT(*) AS [Trading Days],
               ROUND(100.0*SQRT(AVG(x*x)-AVG(x)*AVG(x))*SQRT(252),1) AS [Annual Volatility %]
        FROM r GROUP BY code ORDER BY 3 DESC LIMIT 25""",
    "Max Drawdown": """
        WITH run AS (SELECT code, adj_close,
              MAX(adj_close) OVER (PARTITION BY code ORDER BY trade_date
                  ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) AS peak FROM daily_prices)
        SELECT r.code AS Code, c.company_name AS Company,
               ROUND(MIN(100.0*(r.adj_close-r.peak)/r.peak),1) AS [Max Drawdown %]
        FROM run r JOIN companies c ON c.code=r.code
        GROUP BY r.code ORDER BY 3 ASC LIMIT 25""",
    "Sector Turnover": """
        SELECT s.sector_name AS Sector,
               ROUND(AVG(p.close*p.volume)/1e6,2) AS [Avg Daily Turnover (A$m)]
        FROM daily_prices p JOIN companies c ON c.code=p.code JOIN sectors s ON s.sector_id=c.sector_id
        WHERE p.volume IS NOT NULL GROUP BY s.sector_name ORDER BY 2 DESC""",
    "Companies": """
        SELECT c.code AS Code, c.company_name AS Company, s.sector_name AS Sector,
               ROUND(c.market_cap/1e9,2) AS [Market Cap (A$b)], c.headquarters AS Headquarters
        FROM companies c JOIN sectors s ON s.sector_id=c.sector_id
        ORDER BY c.market_cap DESC""",
}


def style(ws, df):
    for ci in range(1, df.shape[1] + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for ci, col in enumerate(df.columns, 1):
        width = max(len(str(col)), *(df[col].astype(str).map(len).tolist() or [0])) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 42)
    # colour-scale the last numeric column
    last = df.columns[-1]
    if pd.api.types.is_numeric_dtype(df[last]):
        L = get_column_letter(df.shape[1])
        ws.conditional_formatting.add(f"{L}2:{L}{df.shape[0]+1}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))


def main():
    con = sqlite3.connect(DB)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        for name, q in SHEETS.items():
            df = pd.read_sql(q, con)
            df.to_excel(xl, sheet_name=name, index=False)
            style(xl.sheets[name], df)
    con.close()
    print(f"Excel report saved: {OUT}")


if __name__ == "__main__":
    main()
